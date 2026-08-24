from selenium.webdriver.common.by import  By
from openpyxl import load_workbook
import time
from conftest import setup_driver


#TC02
#driver initiation will happen from conftest file
def test_invalid_login(setup_driver):
    def enter_text(driver,locater,value):
        #Keyword driven is implemented to enter the text and click the button
        #this function is created to enter username and password
        locater_type,locater_value=locater.split("=",1)  #since it is having attribute and its value doing split here
    #based on split for 1 time
        if locater_type=="id":
           driver.find_element(By.ID,locater_value).send_keys(value) #Value column from excel sheet
        elif locater_type=='name':
           driver.find_element(By.NAME,locater_value).send_keys(value)
        elif locater_type == 'xpath':
           driver.find_element(By.XPATH, locater_value).send_keys(value)

    def click(driver,locater):
        # this function is created to click on login button
        locater_type, locater_value = locater.split("=", 1)
        if locater_type == "id":
            driver.find_element(By.ID, locater_value).click()
        elif locater_type == 'name':
            driver.find_element(By.NAME, locater_value).click()
        elif locater_type == 'xpath':
            driver.find_element(By.XPATH, locater_value).click()

    def error(driver,locater):
        # this function is created to validate error message
        locater_type, locater_value = locater.split("=", 1)  #here we are splitting the xpath which we are giving
        if locater_type == "id":
            return driver.find_element(By.ID, locater_value)  #here we are returning the error message that we got
        elif locater_type == 'name':
            return driver.find_element(By.NAME, locater_value)
        elif locater_type == 'xpath':
            return driver.find_element(By.XPATH, locater_value)

    def wait(seconds):
        time.sleep(seconds)

    #path of the data sheet
    file_path="D://Malathi ID Proofs/Guvi/KeywordDrivenFP3.xlsx"

    wb=load_workbook(file_path)
     #open the sheet, enter the sheet name inside the bracket
    sheet=wb["Sheet1"] # this is the sheet created for keywords
    data=wb["Sheet2"]   # this is the sheet created for test data
    data_headers = []   #created empty list
    for cell in data[1]:    #it will look at te data sheet for the first row
        if cell.value is not None:  #if it is not empty, it will store the data into the empty list that we created
            data_headers.append(str(cell.value).strip()) #strip is used here to remove the spaces

    for rows in data.iter_rows(min_row=2,values_only=True):     #this is outer loop for data sheet,
        if not rows or rows[0] is None:     #it will continue the execution when the first row of the excel is not empty
            continue

        setup_driver.refresh()      #this will refresh the page for each user

        for row in sheet.iter_rows(min_row=2,values_only=True):     #this is the inner loop for keywords sheet
            if not row or row[0] is None:
                continue

            keyword = row[0]
            locater = row[1]
            value = row[2]

            for idx, header_name in enumerate(data_headers): #here it is fetching details from the data headers list that we created already
                placeholder = f"${{{header_name}}}" #out headername will be assigned to variable placeholder
                if placeholder in str(value):   #it will check the same headername in the keyword sheet(value column)
                    value = str(value).replace(placeholder, str(rows[idx])) #here it will change the header name to actual usrname


            if keyword=="enter_text":   # calling the function
                enter_text(setup_driver,locater,value)
            elif keyword=="click":
                click(setup_driver,locater)
            elif keyword=="wait":
                wait(value)
            elif keyword=="error":

                error_locater=error(setup_driver,locater)
                actual_message=error_locater.text       #fetching the error message text

                clean_actual = " ".join(actual_message.split()).lower() # removing space and making all to lowercase for actual message we are getting in the browser
                clean_expected = " ".join(str(value).split()).lower()# removing space and making all to lowercase for the message that we added in the sheet

                if clean_actual in clean_expected: #comparing both and validating whether both are same or not
                    print("Same error message appeared")
                else:
                    print("Error message is not matched")





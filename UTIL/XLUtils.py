import datetime
import openpyxl


def get_row_count(file,sheet_name):
    workbook= openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file,sheet_name):
    workbook= openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file,sheet_name, row_num, column_num):
    workbook= openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_num).value



def write_data(file,sheet_name, row_num, column_num, data, date_column,date_data,tester_column, tester_data):
    workbook= openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_num, value=data)
    sheet.cell(row=row_num, column=date_column, value=date_data)
    sheet.cell(row=row_num, column=tester_column, value=tester_data)




    workbook.save(file)
    workbook.close()

